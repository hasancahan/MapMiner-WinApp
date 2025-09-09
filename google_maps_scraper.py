import time
import pandas as pd
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import logging

class GoogleMapsScraper:
    def __init__(self, headless=False):
        """
        Google Maps scraper sınıfı
        
        Args:
            headless (bool): Tarayıcıyı görünmez modda çalıştır
        """
        self.driver = None
        self.headless = headless
        self.business_data = []
        self.scan_count = 0  # Tarama sayacı
        self.current_query = ""  # Mevcut arama terimi
        self.current_location = ""  # Mevcut konum
        self.progress_callback = None  # İlerleme callback fonksiyonu
        self.max_results = 0  # Maksimum sonuç sayısı
        
        # Logging ayarları
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        self.logger = logging.getLogger(__name__)
        
    def setup_driver(self):
        """Chrome driver'ı ayarla"""
        try:
            chrome_options = Options()
            if self.headless:
                chrome_options.add_argument("--headless")
            
            # Bot tespitini önlemek için ek ayarlar
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-blink-features=AutomationControlled")
            chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
            chrome_options.add_experimental_option('useAutomationExtension', False)
            chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
            
            self.logger.info("Chrome driver başarıyla ayarlandı")
            return True
            
        except Exception as e:
            self.logger.error(f"Driver ayarlanırken hata: {e}")
            return False
    
    def search_businesses(self, query, location="", max_results=50, detailed_info=True, progress_callback=None):
        """
        Google Maps'te işletme ara - DETAYLI MOD
        
        Args:
            query (str): Aranacak işletme türü (örn: "restoran", "eczane")
            location (str): Konum (örn: "İstanbul", "Ankara")
            max_results (int): Maksimum sonuç sayısı
            detailed_info (bool): Detaylı bilgileri topla (her zaman True)
            progress_callback (function): İlerleme güncelleme callback fonksiyonu
        """
        # Arama bilgilerini sakla
        self.current_query = query
        self.current_location = location
        self.progress_callback = progress_callback
        self.max_results = max_results
        
        # Driver yeniden başlatma limiti kaldırıldı - sınırsız tarama
        # Sadece driver yoksa yeni bir tane oluştur
        
        if not self.driver:
            if not self.setup_driver():
                return False
        
        try:
            # Scan count artık limit için kullanılmıyor - sınırsız tarama
            
            # Google Maps'i aç
            search_query = f"{query} {location}".strip()
            maps_url = f"https://www.google.com/maps/search/{search_query.replace(' ', '+')}"
            
            self.logger.info(f"Arama yapılıyor: {search_query}")
            self.driver.get(maps_url)
            
            # Sayfanın yüklenmesini bekle (daha uzun)
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "[role='main']"))
            )
            
            # Sonuçları topla - detaylı mod
            self._collect_business_data(max_results, detailed_info)
            
            return True
            
        except Exception as e:
            self.logger.error(f"Arama sırasında hata: {e}")
            return False
    
    def _collect_business_data(self, max_results, detailed_info=True):
        """İşletme verilerini topla"""
        collected = 0
        last_height = 0
        scroll_attempts = 0
        max_scroll_attempts = 100  # Artırıldı - sınırsız tarama için
        
        # Sayfanın tam yüklenmesini bekle - detaylı mod
        time.sleep(4)
        
        while collected < max_results and scroll_attempts < max_scroll_attempts:
            try:
                # Farklı selector'ları dene
                business_cards = []
                
                # Yeni Google Maps selector'ları - genişletilmiş
                selectors = [
                    "[data-result-index]",
                    "[jsaction*='pane.resultSection.click']", 
                    "[role='article']",
                    ".Nv2PK",
                    ".THOPZb",
                    ".lI9IFe",
                    "[data-value='Business name']",
                    ".fontHeadlineSmall",
                    ".qBF1Pd",
                    # Ek selector'lar
                    "[data-result-index]",
                    ".VkpGBb",
                    ".Nv2PK.THOPZb",
                    ".lI9IFe.THOPZb",
                    "[jsaction*='pane.resultSection.click']",
                    ".fontHeadlineSmall",
                    ".qBF1Pd.fontHeadlineSmall",
                    # Daha genel selector'lar
                    "[role='button'][jsaction*='pane']",
                    ".fontHeadlineSmall[role='button']",
                    ".qBF1Pd[role='button']"
                ]
                
                for selector in selectors:
                    try:
                        cards = self.driver.find_elements(By.CSS_SELECTOR, selector)
                        if cards:
                            business_cards = cards
                            self.logger.info(f"Selector '{selector}' ile {len(cards)} kart bulundu")
                            break
                    except:
                        continue
                
                # Eğer hiç kart bulunamadıysa, sayfayı kaydır ve tekrar dene
                if not business_cards:
                    self.logger.info("İşletme kartı bulunamadı, sayfa kaydırılıyor...")
                    self._scroll_page()
                    time.sleep(4)  # Daha uzun bekle
                    scroll_attempts += 1
                    continue
                
                self.logger.info(f"Bulunan işletme kartı sayısı: {len(business_cards)}")
                
                # Her işletme kartını işle
                for i, card in enumerate(business_cards[collected:], collected):
                    if collected >= max_results:
                        break
                    
                    try:
                        business_info = self._extract_business_info(card, i)
                        if business_info and business_info.get('Ad'):
                            # Detaylı bilgileri topla - her zaman detaylı mod
                            try:
                                detailed_info_data = self.get_detailed_info(i)
                                business_info.update(detailed_info_data)
                                self.logger.info(f"Detaylı bilgiler toplandı: {business_info['Ad']}")
                            except Exception as e:
                                self.logger.warning(f"Detaylı bilgi toplama hatası: {e}")
                            
                            self.business_data.append(business_info)
                            collected += 1
                            self.logger.info(f"Toplanan işletme sayısı: {collected} - {business_info['Ad']}")
                            
                            # İlerleme çubuğunu güncelle
                            if self.progress_callback:
                                progress_percent = (collected / self.max_results) * 100
                                self.progress_callback(progress_percent, collected, business_info['Ad'])
                            
                            # 10 tane bulunca özel bekleme ve scroll
                            if collected == 10:
                                self.logger.info("10 işletme bulundu, özel scroll yapılıyor...")
                                time.sleep(1)  # 1 saniye bekle
                                
                                # 1 kez scroll yap
                                self.logger.info("Özel scroll yapılıyor...")
                                self._scroll_page()
                                time.sleep(1)
                                
                                self.logger.info("Özel scroll tamamlandı, devam ediliyor...")
                            
                    except Exception as e:
                        self.logger.warning(f"İşletme bilgisi çıkarılırken hata: {e}")
                        continue
                
                # Sayfayı aşağı kaydır
                if collected < max_results:
                    self._scroll_page()
                    # Detaylı mod için bekleme süresi
                    time.sleep(4)
                    scroll_attempts += 1
                    
                    # Yeni içerik yüklenip yüklenmediğini kontrol et
                    new_height = self.driver.execute_script("return document.body.scrollHeight")
                    if new_height == last_height:
                        # Daha agresif kaydırma dene
                        self.logger.info("Sayfa yüksekliği değişmedi, daha agresif kaydırma deneniyor...")
                        
                        # Daha fazla kez kaydır - sayfa sonlarında daha derinlemesine
                        for _ in range(5):  # 3'ten 5'e çıkarıldı
                            self._scroll_page()
                            time.sleep(2)
                            
                            # Ekstra derinlemesine kaydırma
                            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight + 1000);")
                            time.sleep(1)
                            
                            new_height = self.driver.execute_script("return document.body.scrollHeight")
                            if new_height != last_height:
                                break
                        
                        # Son bir kez daha agresif kaydırma
                        if new_height == last_height:
                            self.logger.info("Son deneme: Çok agresif kaydırma...")
                            for _ in range(3):
                                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight + 2000);")
                                time.sleep(2)
                                self._scroll_page()
                                time.sleep(2)
                                new_height = self.driver.execute_script("return document.body.scrollHeight")
                                if new_height != last_height:
                                    break
                        
                    if new_height == last_height:
                        self.logger.info("Daha fazla sonuç bulunamadı")
                        break
                    last_height = new_height
                
            except Exception as e:
                self.logger.error(f"Veri toplama sırasında hata: {e}")
                break
        
        # İstenilen sayıya ulaşıldı mı kontrol et
        if len(self.business_data) >= max_results:
            self.logger.info(f"🎉 HEDEF ULAŞILDI! {len(self.business_data)} işletme bulundu (Hedef: {max_results})")
            self._show_completion_notification(len(self.business_data), max_results)
            self._print_scan_report()
        else:
            self.logger.info(f"Toplam {len(self.business_data)} işletme bulundu (Hedef: {max_results})")
    
    def _show_completion_notification(self, found_count, target_count):
        """Tamamlanma bildirimi göster"""
        try:
            import tkinter as tk
            from tkinter import messagebox
            
            # Basit bildirim penceresi
            root = tk.Tk()
            root.withdraw()  # Ana pencereyi gizle
            
            messagebox.showinfo(
                "🎉 Tarama Tamamlandı!",
                f"Hedef: {target_count} işletme\n"
                f"Bulunan: {found_count} işletme\n\n"
                f"Tarama başarıyla tamamlandı!"
            )
            
            root.destroy()
        except Exception as e:
            self.logger.debug(f"Bildirim gösterilemedi: {e}")
    
    def _print_scan_report(self):
        """Konsol tarama raporu yazdır"""
        try:
            print("\n" + "="*60)
            print("📊 TARAMA RAPORU")
            print("="*60)
            print(f"📅 Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
            print(f"🎯 Hedef: {len(self.business_data)} işletme")
            print(f"✅ Bulunan: {len(self.business_data)} işletme")
            print(f"📈 Başarı Oranı: %100")
            print("-"*60)
            
            # İstatistikler
            with_phone = sum(1 for b in self.business_data if b.get('Telefon'))
            with_address = sum(1 for b in self.business_data if b.get('Adres'))
            with_rating = sum(1 for b in self.business_data if b.get('Puan/Yorum'))
            
            print(f"📞 Telefon: {with_phone}/{len(self.business_data)} (%{with_phone/len(self.business_data)*100:.1f})")
            print(f"📍 Adres: {with_address}/{len(self.business_data)} (%{with_address/len(self.business_data)*100:.1f})")
            print(f"⭐ Puan: {with_rating}/{len(self.business_data)} (%{with_rating/len(self.business_data)*100:.1f})")
            print("-"*60)
            
            # İlk 5 sonuç
            print("🏆 İLK 5 SONUÇ:")
            for i, business in enumerate(self.business_data[:5], 1):
                print(f"{i}. {business.get('Ad', 'Bilinmiyor')}")
                if business.get('Adres'):
                    print(f"   📍 {business['Adres']}")
                if business.get('Telefon'):
                    print(f"   📞 {business['Telefon']}")
                print()
            
            print("="*60)
            
        except Exception as e:
            self.logger.error(f"Rapor yazdırılamadı: {e}")

    def _extract_business_info(self, card, index):
        """İşletme kartından bilgileri çıkar"""
        try:
            business_info = {
                'Sıra': index + 1,
                'Ad': '',
                'Adres': '',
                'Telefon': '',
                'Puan/Yorum': ''
            }
            
            # İşletme adını al - farklı selector'ları dene
            name_selectors = [
                "[data-value='Business name']",
                "h3", "h2", "[role='heading']",
                ".fontHeadlineSmall",
                ".qBF1Pd",
                ".fontDisplayLarge",
                ".fontHeadlineMedium"
            ]
            
            for selector in name_selectors:
                try:
                    name_element = card.find_element(By.CSS_SELECTOR, selector)
                    if name_element.text.strip():
                        business_info['Ad'] = name_element.text.strip()
                        break
                except:
                    continue
            
            # Ana sayfada telefon ve adres genellikle yok, sadece puan var
            # Bu bilgiler detaylı sayfada bulunur
            
            # Website bilgisini al
            try:
                website_element = card.find_element(By.CSS_SELECTOR, "[data-item-id='authority']")
                business_info['Website'] = website_element.text.strip()
            except:
                pass
            
            # Puan ve yorum sayısını al (yeni yöntem)
            self._extract_rating_new_way(card, business_info)
            
            # Kategori bilgisini al
            try:
                category_element = card.find_element(By.CSS_SELECTOR, "[data-value='Category']")
                business_info['Kategori'] = category_element.text.strip()
            except:
                pass
            
            # Açılış saatleri
            try:
                hours_element = card.find_element(By.CSS_SELECTOR, "[data-item-id='oh']")
                business_info['Açılış Saatleri'] = hours_element.text.strip()
            except:
                pass
            
            # Durum (açık/kapalı)
            try:
                status_element = card.find_element(By.CSS_SELECTOR, "[data-value='Open hours status']")
                business_info['Durum'] = status_element.text.strip()
            except:
                pass
            
            # Fiyat seviyesi
            try:
                price_element = card.find_element(By.CSS_SELECTOR, "[data-value='Price']")
                business_info['Fiyat Seviyesi'] = price_element.text.strip()
            except:
                # Alternatif fiyat selector'ları
                price_selectors = [".fontBodyMedium", ".fontBodySmall"]
                for selector in price_selectors:
                    try:
                        elements = card.find_elements(By.CSS_SELECTOR, selector)
                        for elem in elements:
                            text = elem.text.strip()
                            if '₺' in text or 'TL' in text or 'price' in text.lower():
                                business_info['Fiyat Seviyesi'] = text
                                break
                        if business_info['Fiyat Seviyesi']:
                            break
                    except:
                        continue
            
            # Detaylı bilgileri ayrı ayrı al
            self._extract_detailed_info_from_card(card, business_info)
            
            # Eğer isim bulunamadıysa None döndür
            if not business_info['Ad']:
                return None
                
            return business_info
            
        except Exception as e:
            self.logger.warning(f"İşletme bilgisi çıkarılırken hata: {e}")
            return None
    
    def _extract_detailed_info_from_card(self, card, business_info):
        """Karttan sadece gerekli bilgileri çıkar - YENİ YAKLAŞIM"""
        try:
            # Yeni Google Maps yapısına uygun selector'lar
            self._extract_phone_new_way(card, business_info)
            self._extract_address_new_way(card, business_info)
            self._extract_rating_new_way(card, business_info)
            
        except Exception as e:
            self.logger.warning(f"Detaylı bilgi çıkarma hatası: {e}")
    
    def _extract_phone_new_way(self, card, business_info):
        """Yeni yöntemle telefon numarası çıkar - GOOGLE MAPS YAPISINA UYGUN"""
        import re
        
        if business_info['Telefon']:  # Zaten telefon var
            return
        
        # Google Maps'in gerçek yapısına uygun selector'lar
        phone_selectors = [
            # Yeni Google Maps yapısı
            "[data-item-id^='phone']",
            "button[aria-label*='Telefon:']",
            "a[href^='tel:']",
            # Text elementleri
            ".Io6YTe.fontBodyMedium"
        ]
        
        for selector in phone_selectors:
            try:
                elements = card.find_elements(By.CSS_SELECTOR, selector)
                for element in elements:
                    # aria-label'den telefon çıkar
                    aria_label = element.get_attribute('aria-label') or ''
                    if 'Telefon:' in aria_label:
                        phone_match = re.search(r'Telefon:\s*([0-9\s\(\)]+)', aria_label)
                        if phone_match:
                            phone = phone_match.group(1).strip()
                            if self._is_valid_phone(phone):
                                business_info['Telefon'] = phone
                                return
                    
                    # href'den telefon çıkar
                    href = element.get_attribute('href') or ''
                    if 'tel:' in href:
                        phone = href.replace('tel:', '').strip()
                        if self._is_valid_phone(phone):
                            business_info['Telefon'] = phone
                            return
                    
                    # Text'ten telefon çıkar
                    text = element.text.strip()
                    if text and self._is_valid_phone(text):
                        business_info['Telefon'] = text
                        return
                        
            except Exception:
                continue
    
    def _extract_address_new_way(self, card, business_info):
        """Yeni yöntemle adres çıkar - GOOGLE MAPS YAPISINA UYGUN"""
        import re
        
        if business_info['Adres']:  # Zaten adres var
            return
        
        # Google Maps'in gerçek yapısına uygun selector'lar
        address_selectors = [
            # Yeni Google Maps yapısı
            "[data-item-id='address']",
            "button[aria-label*='Adres:']",
            # Text elementleri
            ".Io6YTe.fontBodyMedium"
        ]
        
        for selector in address_selectors:
            try:
                elements = card.find_elements(By.CSS_SELECTOR, selector)
                for element in elements:
                    # aria-label'den adres çıkar
                    aria_label = element.get_attribute('aria-label') or ''
                    if 'Adres:' in aria_label:
                        address_match = re.search(r'Adres:\s*(.+)', aria_label)
                        if address_match:
                            address = address_match.group(1).strip()
                            if self._is_clean_address(address):
                                business_info['Adres'] = address
                                return
                    
                    # Text'ten adres çıkar
                    text = element.text.strip()
                    if text and self._is_clean_address(text):
                        business_info['Adres'] = text
                        return
                        
            except Exception:
                continue
    
    def _extract_rating_new_way(self, card, business_info):
        """Yeni yöntemle puan çıkar"""
        import re
        
        if business_info['Puan/Yorum']:  # Zaten puan var
            return
        
        # Google Maps'in yeni yapısına uygun selector'lar
        rating_selectors = [
            # Yeni Google Maps yapısı
            "[role='img']",
            "[aria-label*='star']",
            "[aria-label*='yıldız']",
            # Text elementleri
            ".fontDisplayMedium",
            ".fontBodySmall",
            # Spesifik puan göstergeleri
            "[data-value*='rating']"
        ]
        
        for selector in rating_selectors:
            try:
                elements = card.find_elements(By.CSS_SELECTOR, selector)
                for element in elements:
                    text = element.text.strip()
                    aria_label = element.get_attribute('aria-label') or ''
                    
                    # aria-label'den puan çıkar
                    if aria_label and ('star' in aria_label.lower() or 'yıldız' in aria_label.lower()):
                        business_info['Puan/Yorum'] = aria_label
                        return
                    
                    # Text'ten puan çıkar
                    if text and re.search(r'\d+,\d+\(\d+\)', text):
                        business_info['Puan/Yorum'] = text
                        return
                        
            except Exception:
                continue
    
    def _is_valid_phone(self, text):
        """Telefon numarası geçerliliğini kontrol et - ESNEK YAKLAŞIM"""
        import re
        
        if not text:
            return False
        
        # Temizle
        clean_text = re.sub(r'[^\d]', '', text)
        
        # En az 7, en fazla 11 haneli olmalı (daha esnek)
        if len(clean_text) < 7 or len(clean_text) > 11:
            return False
        
        # Türkiye telefon numarası formatları (daha esnek)
        phone_patterns = [
            r'\(\d{3,4}\)\s*\d{3,4}\s*\d{2,4}',  # (0212) 123 45 67
            r'0\d{3,4}\s*\d{3,4}\s*\d{2,4}',      # 0212 123 45 67
            r'\d{3,4}\s*\d{3,4}\s*\d{2,4}',       # 212 123 45 67
            r'05\d{2}\s*\d{3}\s*\d{2}\s*\d{2}',   # 0543 823 00 00
            r'5\d{2}\s*\d{3}\s*\d{2}\s*\d{2}',    # 543 823 00 00
            r'\+90\s*\d{3,4}\s*\d{3,4}\s*\d{2,4}', # +90 212 123 45 67
            r'90\s*\d{3,4}\s*\d{3,4}\s*\d{2,4}',   # 90 212 123 45 67
            r'\d{3,4}\s*\d{3,4}\s*\d{2,4}',       # 212 123 45 67
            r'\d{3,4}\s*\d{3,4}',                  # 212 123 45
            r'\d{3,4}\s*\d{3,4}\s*\d{2}',         # 212 123 45
        ]
        
        for pattern in phone_patterns:
            if re.search(pattern, text):
                return True
        
        return False
    
    def _is_valid_address(self, text):
        """Adres geçerliliğini kontrol et - ESNEK YAKLAŞIM"""
        import re
        
        if not text:
            return False
        
        # Adres olmayan metinleri filtrele
        invalid_patterns = [
            r'^\d+,\d+\(\d+\)$',  # Puan bilgisi
            r'^\d+,\d+\(\d+\)',   # Puan bilgisi ile başlayan
            r'^\d+$',             # Sadece sayı
            r'^[A-Za-z\s]{1,10}$', # Çok kısa sadece harf
        ]
        
        for pattern in invalid_patterns:
            if re.search(pattern, text):
                return False
        
        # Geçerli adres göstergeleri
        address_indicators = [
            'sokak', 'sk', 'cadde', 'cd', 'bulvar', 'blv', 'mahalle', 'mah',
            'no:', 'no ', 'apt', 'daire', 'kat', 'sok', 'cad', 'blv',
            'taksi', 'durak', 'durağı', 'merkez', 'plaza', 'avm', 'center'
        ]
        
        text_lower = text.lower()
        for indicator in address_indicators:
            if indicator in text_lower:
                return True
        
        # Uzun metinler adres olabilir (daha esnek)
        if len(text) > 10:
            return True
        
        return False
    
    def _is_clean_address(self, text):
        """Temiz adres kontrolü - sadece gerçek adres metinleri"""
        import re
        
        if not text:
            return False
        
        # Adres olmayan metinleri filtrele
        invalid_patterns = [
            r'^\d+,\d+\(\d+\)$',  # Puan bilgisi
            r'^\d+,\d+\(\d+\)',   # Puan bilgisi ile başlayan
            r'^\d+$',             # Sadece sayı
            r'^[A-Za-z\s]{1,10}$', # Çok kısa sadece harf
            r'\(0\d{3,4}\)',      # Telefon numarası
            r'0\d{3,4}\s*\d{3,4}', # Telefon numarası
            r'Web sitesi',        # Web sitesi
            r'Yol tarifi',        # Yol tarifi
            r'Şoförleri',         # Yorum metni
            r'Deneyimli',         # Yorum metni
            r'Günün',             # Yorum metni
            r'Taksi durağı',      # Genel ifade
        ]
        
        for pattern in invalid_patterns:
            if re.search(pattern, text):
                return False
        
        # Geçerli adres göstergeleri
        address_indicators = [
            'sokak', 'sk', 'cadde', 'cd', 'bulvar', 'blv', 'mahalle', 'mah',
            'no:', 'no ', 'apt', 'daire', 'kat', 'sok', 'cad', 'blv',
            'profesör', 'doktor', 'caddesi', 'mimoza', 'meşrutiyet',
            'paşa', 'atif', 'yılmaz', 'yüzyıl', 'galericiler'
        ]
        
        text_lower = text.lower()
        for indicator in address_indicators:
            if indicator in text_lower:
                return True
        
        # Uzun metinler adres olabilir (daha esnek)
        if len(text) > 20:
            return True
        
        return False
    
    def _parse_combined_text(self, text, business_info):
        """Birleşik metni sadece gerekli bilgilere böl"""
        import re
        
        lines = text.split('\n')
        
        # Gelişmiş telefon numarası çıkarma
        self._extract_phone_numbers(text, business_info)
        
        # Puan/Yorum bilgisini çıkar (tüm metinden)
        rating_match = re.search(r'(\d+,\d+\(\d+\))', text)
        if rating_match and not business_info['Puan/Yorum']:
            business_info['Puan/Yorum'] = rating_match.group(1)
        
        # Her satırı analiz et
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Puan bilgisini atla (zaten var)
            if re.match(r'^\d+,\d+\(\d+\)$', line):
                continue
            
            # Adres çıkar (uzun satırlar ve adres göstergeleri)
            if self._is_address(line) and not business_info['Adres']:
                business_info['Adres'] = line
                continue
        
        # Eğer adres bulunamadıysa, en uzun satırı adres olarak al (puan bilgisi hariç)
        if not business_info['Adres']:
            longest_line = max(lines, key=len) if lines else ""
            if (len(longest_line) > 20 and 
                not self._is_phone_number(longest_line) and 
                not re.match(r'^\d+,\d+\(\d+\)$', longest_line) and
                not re.search(r'\d+,\d+\(\d+\)', longest_line)):
                business_info['Adres'] = longest_line
    
    def _extract_phone_numbers(self, text, business_info):
        """Gelişmiş telefon numarası çıkarma"""
        import re
        
        if business_info['Telefon']:  # Zaten telefon var
            return
        
        # Türkiye telefon numarası formatları
        phone_patterns = [
            # (0212) 123 45 67
            r'\(\d{3,4}\)\s*\d{3,4}\s*\d{2,4}',
            # 0212 123 45 67
            r'0\d{3,4}\s*\d{3,4}\s*\d{2,4}',
            # 212 123 45 67
            r'\d{3,4}\s*\d{3,4}\s*\d{2,4}',
            # 0543 823 00 00
            r'05\d{2}\s*\d{3}\s*\d{2}\s*\d{2}',
            # 543 823 00 00
            r'5\d{2}\s*\d{3}\s*\d{2}\s*\d{2}',
            # +90 212 123 45 67
            r'\+90\s*\d{3,4}\s*\d{3,4}\s*\d{2,4}',
            # 90 212 123 45 67
            r'90\s*\d{3,4}\s*\d{3,4}\s*\d{2,4}',
            # 212-123-45-67
            r'\d{3,4}-\d{3,4}-\d{2,4}',
            # 212.123.45.67
            r'\d{3,4}\.\d{3,4}\.\d{2,4}',
            # 2121234567 (10 haneli)
            r'\b\d{10}\b',
            # 212123456 (9 haneli)
            r'\b\d{9}\b'
        ]
        
        found_phones = []
        
        for pattern in phone_patterns:
            matches = re.findall(pattern, text)
            for match in matches:
                # Temizle ve kontrol et
                clean_phone = re.sub(r'[^\d]', '', match)
                if len(clean_phone) >= 9:  # En az 9 haneli olmalı
                    found_phones.append(match.strip())
        
        if found_phones:
            # En uzun ve en geçerli telefon numarasını seç
            best_phone = max(found_phones, key=len)
            business_info['Telefon'] = best_phone
    
    def _is_phone_number(self, text):
        """Telefon numarası kontrolü"""
        phone_patterns = [
            r'\(\d{3,4}\)\s*\d{3,4}\s*\d{2,4}',  # (0212) 123 45 67
            r'\d{3,4}\s*\d{3,4}\s*\d{2,4}',      # 0212 123 45 67
            r'\+90\s*\d{3,4}\s*\d{3,4}\s*\d{2,4}', # +90 212 123 45 67
        ]
        import re
        for pattern in phone_patterns:
            if re.search(pattern, text):
                return True
        return False
    
    def _is_address(self, text):
        """Adres kontrolü"""
        address_indicators = ['Mah', 'Cad', 'Sok', 'No:', 'Blok', 'Kat', 'Daire', 'Mahallesi', 'Caddesi', 'Sokağı', 'Sk.', 'Cd.', 'Apt.', 'Sitesi']
        return any(indicator in text for indicator in address_indicators) and len(text) > 15
    
    def _is_category(self, text):
        """Kategori kontrolü"""
        category_indicators = [
            'Taksi', 'Eczane', 'Restoran', 'Market', 'Cafe', 'Otel', 'Hastane', 
            'Okul', 'Banka', 'Emlak', 'Berber', 'Kuaför', 'Spor', 'Fitness',
            'Durağı', 'Merkezi', 'Şubesi', 'Mağazası', 'Dükkanı', 'Çilingir',
            'Anahtarcı', 'Hizmetleri', 'Servisi', 'Elektronik', 'Oto'
        ]
        return any(indicator in text for indicator in category_indicators) and len(text) < 60
    
    def _is_website(self, text):
        """Website kontrolü"""
        website_indicators = ['www.', 'http', '.com', '.tr', '.org', '.net']
        return any(indicator in text.lower() for indicator in website_indicators)
    
    def _is_opening_hours(self, text):
        """Açılış saatleri kontrolü"""
        time_patterns = [
            r'\d{1,2}:\d{2}',  # 09:00
            r'\d{1,2}\.\d{2}',  # 09.00
            r'Pazartesi|Salı|Çarşamba|Perşembe|Cuma|Cumartesi|Pazar',
            r'Kapalı|Açık|Hafta|Gün'
        ]
        import re
        for pattern in time_patterns:
            if re.search(pattern, text, re.IGNORECASE):
                return True
        return False
    
    def _is_status(self, text):
        """Durum kontrolü"""
        status_indicators = ['Açık', 'Kapalı', 'Open', 'Closed', 'Şu anda açık', 'Şu anda kapalı']
        return any(indicator in text for indicator in status_indicators)
    
    def _scroll_page(self):
        """Sayfayı aşağı kaydır - GELİŞTİRİLMİŞ"""
        try:
            # Google Maps'in yeni yapısı için gelişmiş kaydırma
            # Önce ana sonuçlar panelini bul
            results_panel = self.driver.find_element(By.CSS_SELECTOR, "[role='main']")
            
            # Panel içinde kaydır
            self.driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", results_panel)
            time.sleep(1)
            
            # Alternatif: Sonuç listesini bul ve kaydır
            try:
                results_list = self.driver.find_element(By.CSS_SELECTOR, "[role='feed']")
                self.driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", results_list)
                time.sleep(1)
            except:
                pass
            
            # Son olarak genel sayfa kaydırma - daha agresif
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight + 500);")
            time.sleep(2)  # Kaydırma sonrası bekle
            
            # Ekstra kaydırma - sayfa sonlarında daha derinlemesine
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight + 1000);")
            time.sleep(1)
            
            # Daha fazla sonuç yüklemek için "Daha fazla göster" butonunu ara
            # 'Daha fazla göster' butonu kaldırıldı - scroll otomatik yükler
                
        except Exception as e:
            self.logger.warning(f"Kaydırma hatası: {e}")
            # Son çare: genel sayfa kaydırma
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
    

    def get_detailed_info(self, business_index):
        """Belirli bir işletmenin detaylı bilgilerini al"""
        try:
            # İşletme kartına tıkla
            business_cards = self.driver.find_elements(By.CSS_SELECTOR, ".Nv2PK")
            if not business_cards:
                business_cards = self.driver.find_elements(By.CSS_SELECTOR, "[data-result-index]")
            
            if business_cards and business_index < len(business_cards):
                business_cards[business_index].click()
                time.sleep(1)  # Çok hızlı bekleme
                
                # Detaylı bilgileri topla
                detailed_info = {}
                
                # TELEFON BİLGİSİNİ AL (Google Maps yapısına uygun)
                try:
                    phone_elements = self.driver.find_elements(By.CSS_SELECTOR, "[data-item-id^='phone']")
                    for element in phone_elements:
                        aria_label = element.get_attribute('aria-label') or ''
                        if 'Telefon:' in aria_label:
                            import re
                            phone_match = re.search(r'Telefon:\s*([0-9\s\(\)]+)', aria_label)
                            if phone_match:
                                detailed_info['Telefon'] = phone_match.group(1).strip()
                                break
                except:
                    pass
                
                # ADRES BİLGİSİNİ AL (Google Maps yapısına uygun)
                try:
                    address_elements = self.driver.find_elements(By.CSS_SELECTOR, "[data-item-id='address']")
                    for element in address_elements:
                        aria_label = element.get_attribute('aria-label') or ''
                        if 'Adres:' in aria_label:
                            import re
                            address_match = re.search(r'Adres:\s*(.+)', aria_label)
                            if address_match:
                                detailed_info['Adres'] = address_match.group(1).strip()
                                break
                except:
                    pass
                
                # Sadece telefon ve adres bilgilerini al, diğer bilgileri kaldır
                
                return detailed_info
                
        except Exception as e:
            self.logger.error(f"Detaylı bilgi alınırken hata: {e}")
            return {}
    
    def generate_filename(self):
        """Otomatik dosya ismi oluştur: İşletme_Şehir_Raporu_Tarih"""
        try:
            from datetime import datetime
            import re
            
            # Tarih formatı: YYYY-MM-DD
            current_date = datetime.now().strftime("%Y-%m-%d")
            
            # Query'yi temizle (özel karakterleri kaldır)
            clean_query = re.sub(r'[^\w\s]', '', self.current_query)
            clean_query = re.sub(r'\s+', '_', clean_query.strip())
            
            # Location'ı temizle
            clean_location = re.sub(r'[^\w\s]', '', self.current_location)
            clean_location = re.sub(r'\s+', '_', clean_location.strip())
            
            # Eğer location boşsa "Genel" kullan
            if not clean_location:
                clean_location = "Genel"
            
            # Dosya ismini oluştur
            filename = f"{clean_query}_{clean_location}_Raporu_{current_date}.xlsx"
            
            return filename
            
        except Exception as e:
            self.logger.error(f"Dosya ismi oluşturulurken hata: {e}")
            # Hata durumunda varsayılan isim
            from datetime import datetime
            current_date = datetime.now().strftime("%Y-%m-%d")
            return f"İşletme_Raporu_{current_date}.xlsx"
    
    def save_to_excel(self, filename=None):
        """Toplanan verileri Excel dosyasına kaydet - MODERN VE RENKLİ"""
        try:
            if not self.business_data:
                self.logger.warning("Kaydedilecek veri bulunamadı")
                return False
            
            # Eğer filename verilmemişse otomatik oluştur
            if filename is None:
                filename = self.generate_filename()
            
            df = pd.DataFrame(self.business_data)
            
            # Sadece belirli sütunları seç
            required_columns = ['Sıra', 'Ad', 'Adres', 'Telefon', 'Puan/Yorum']
            available_columns = [col for col in required_columns if col in df.columns]
            df_filtered = df[available_columns]
            
            # Excel dosyasını oluştur
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df_filtered.to_excel(writer, sheet_name='İşletme Verileri', index=False)
                
                # Worksheet'i al
                worksheet = writer.sheets['İşletme Verileri']
                
                # Modern stil için gerekli import'lar
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                from openpyxl.utils import get_column_letter
                from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
                from openpyxl.styles.differential import DifferentialStyle
                
                # Sayfa başlığını EN BAŞTA ekle (stil uygulamalarından önce)
                worksheet.insert_rows(1)
                title_cell = worksheet.cell(row=1, column=1)
                title_cell.value = f"🗺️ GOOGLE MAPS İŞLETME VERİLERİ - {len(df_filtered)} İşletme"
                title_cell.font = Font(name='Calibri', size=16, bold=True, color='6F42C1')
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Başlık satırını birleştir
                worksheet.merge_cells(f'A1:{get_column_letter(len(df_filtered.columns))}1')
                
                # Renk paleti
                colors = {
                    'header_bg': '2E86AB',      # Mavi
                    'header_text': 'FFFFFF',    # Beyaz
                    'row_light': 'F8F9FA',      # Açık gri
                    'row_dark': 'E9ECEF',       # Koyu gri
                    'border': 'DEE2E6',         # Kenarlık
                    'high_rating': '28A745',    # Yeşil (yüksek puan)
                    'medium_rating': 'FFC107',  # Sarı (orta puan)
                    'low_rating': 'DC3545',     # Kırmızı (düşük puan)
                    'accent': '6F42C1'          # Mor (vurgu)
                }
                
                # Kenarlık stili
                thin_border = Border(
                    left=Side(style='thin', color=colors['border']),
                    right=Side(style='thin', color=colors['border']),
                    top=Side(style='thin', color=colors['border']),
                    bottom=Side(style='thin', color=colors['border'])
                )
                
                # Başlık stili
                header_font = Font(
                    name='Calibri',
                    size=12,
                    bold=True,
                    color=colors['header_text']
                )
                header_fill = PatternFill(
                    start_color=colors['header_bg'],
                    end_color=colors['header_bg'],
                    fill_type='solid'
                )
                
                # Veri stili
                data_font = Font(name='Calibri', size=11)
                center_alignment = Alignment(horizontal='center', vertical='center')
                left_alignment = Alignment(horizontal='left', vertical='center')
                
                # Başlık satırını stilleyin (2. satır) - sadece gerekli sütunlar
                for col_num in range(1, len(df_filtered.columns) + 1):
                    cell = worksheet.cell(row=2, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = center_alignment
                
                # Veri satırlarını stilleyin - tüm satırlar dahil
                total_rows = len(df_filtered) + 2  # +1 başlık satırı +1 title satırı için
                for row_num in range(3, total_rows + 1):
                    for col_num in range(1, len(df_filtered.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.font = data_font
                        cell.border = thin_border
                        
                        # Alternatif satır renkleri - tüm satırlar için
                        if row_num % 2 == 0:
                            cell.fill = PatternFill(
                                start_color=colors['row_light'],
                                end_color=colors['row_light'],
                                fill_type='solid'
                            )
                        else:
                            cell.fill = PatternFill(
                                start_color=colors['row_dark'],
                                end_color=colors['row_dark'],
                                fill_type='solid'
                            )
                        
                        # Sütun hizalaması
                        if col_num == 1:  # Sıra sütunu
                            cell.alignment = center_alignment
                        elif col_num == 5:  # Puan/Yorum sütunu
                            cell.alignment = center_alignment
                        else:
                            cell.alignment = left_alignment
                
                # Sütun genişliklerini optimize et - sadece gerekli sütunlar
                column_widths = {
                    'A': 8,   # Sıra
                    'B': 45,  # Ad
                    'C': 60,  # Adres
                    'D': 25,  # Telefon
                    'E': 18   # Puan/Yorum
                }
                
                for col_letter, width in column_widths.items():
                    if col_letter in [get_column_letter(i) for i in range(1, len(df_filtered.columns) + 1)]:
                        worksheet.column_dimensions[col_letter].width = width
                
                # Puan sütunu için koşullu formatlama
                if 'Puan/Yorum' in df_filtered.columns:
                    puan_col = df_filtered.columns.get_loc('Puan/Yorum') + 1
                    
                    # Puan verilerini temizle ve sayısal değere çevir - tüm satırlar için
                    for row_num in range(3, total_rows + 1):
                        cell = worksheet.cell(row=row_num, column=puan_col)
                        if cell.value:
                            # Puan verisini temizle (örn: "4,8 yıldızlı 4.142 Yorum" -> 4.8)
                            import re
                            rating_match = re.search(r'(\d+,\d+)', str(cell.value))
                            if rating_match:
                                rating = float(rating_match.group(1).replace(',', '.'))
                                cell.value = rating
                    
                    # Yüksek puan (4.5+) - Yeşil
                    high_rating_rule = CellIsRule(
                        operator='greaterThan',
                        formula=['4.4'],
                        fill=PatternFill(
                            start_color=colors['high_rating'],
                            end_color=colors['high_rating'],
                            fill_type='solid'
                        ),
                        font=Font(color='FFFFFF', bold=True)
                    )
                    
                    # Orta puan (3.5-4.4) - Sarı
                    medium_rating_rule = CellIsRule(
                        operator='between',
                        formula=['3.4', '4.4'],
                        fill=PatternFill(
                            start_color=colors['medium_rating'],
                            end_color=colors['medium_rating'],
                            fill_type='solid'
                        ),
                        font=Font(color='000000', bold=True)
                    )
                    
                    # Düşük puan (<3.5) - Kırmızı
                    low_rating_rule = CellIsRule(
                        operator='lessThan',
                        formula=['3.5'],
                        fill=PatternFill(
                            start_color=colors['low_rating'],
                            end_color=colors['low_rating'],
                            fill_type='solid'
                        ),
                        font=Font(color='FFFFFF', bold=True)
                    )
                    
                    # Koşullu formatlamayı uygula - tüm satırlar için
                    worksheet.conditional_formatting.add(
                        f'{get_column_letter(puan_col)}3:{get_column_letter(puan_col)}{total_rows}',
                        high_rating_rule
                    )
                    worksheet.conditional_formatting.add(
                        f'{get_column_letter(puan_col)}3:{get_column_letter(puan_col)}{total_rows}',
                        medium_rating_rule
                    )
                    worksheet.conditional_formatting.add(
                        f'{get_column_letter(puan_col)}3:{get_column_letter(puan_col)}{total_rows}',
                        low_rating_rule
                    )
                
                # Satır yüksekliğini ayarla - tüm satırlar için
                for row_num in range(1, total_rows + 1):
                    worksheet.row_dimensions[row_num].height = 25
                
                # Başlık satırı (1. satır) için stil
                for col_num in range(1, len(df_filtered.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = PatternFill(
                        start_color='F8F9FA',
                        end_color='F8F9FA',
                        fill_type='solid'
                    )
                    cell.border = thin_border
            
            self.logger.info(f"🎨 Modern ve renkli Excel dosyası başarıyla oluşturuldu: {filename}")
            self.logger.info(f"📊 Toplam {len(self.business_data)} işletme verisi kaydedildi")
            return True
            
        except Exception as e:
            self.logger.error(f"Excel dosyası kaydedilirken hata: {e}")
            return False
    
    def close(self):
        """Driver'ı kapat"""
        if self.driver:
            self.driver.quit()
            self.logger.info("Driver kapatıldı")
    
    def __enter__(self):
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
